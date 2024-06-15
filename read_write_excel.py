import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook


def get_all_products():
    # Give the location of the file
    path = "ExcelData.xlsx"
    
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row

    data_dict = {
        'Products': [],
        'Category': [],
        'Date': [],
        'Time of subscription': [],
        'Price in shop': [],
        'Price of bought': [],
        'Profit': []
    }

    # Loop through all rows starting from row 2
    for i in range(2, max_row + 1):
        data_dict['Products'].append(sheet_obj.cell(row=i, column=1).value)
        data_dict['Category'].append(sheet_obj.cell(row=i, column=2).value)
        data_dict['Date'].append(sheet_obj.cell(row=i, column=3).value)
        data_dict['Time of subscription'].append(sheet_obj.cell(row=i, column=4).value)
        data_dict['Price in shop'].append(sheet_obj.cell(row=i, column=5).value)
        data_dict['Price of bought'].append(sheet_obj.cell(row=i, column=6).value)
        data_dict['Profit'].append(sheet_obj.cell(row=i, column=7).value)

    return data_dict

def add_product_row(Product,Category,Date,Time_of_subscription,Price_in_shop,Price_of_bought,Profit):
     # Give the location of the file
    path = "ExcelData.xlsx"
    
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row

    #max_row is a sheet function that gets the last row in a sheet.
    newRowLocation = sheet_obj.max_row +1

    #write to the cell you want, specifying row and column, and value :
    sheet_obj.cell(column=1,row=newRowLocation, value=str(Product))
    sheet_obj.cell(column=2,row=newRowLocation, value=str(Category))
    sheet_obj.cell(column=3,row=newRowLocation, value=str(Date))
    sheet_obj.cell(column=4,row=newRowLocation, value=str(Time_of_subscription))
    sheet_obj.cell(column=5,row=newRowLocation, value=round(float(Price_in_shop),3))
    sheet_obj.cell(column=6,row=newRowLocation, value=round(float(Price_of_bought),3))
    sheet_obj.cell(column=7,row=newRowLocation, value=round(float(Profit),3))

    wb_obj.save(filename=path)
    wb_obj.close()





# # Print each column data
# for key, value in data_dict.items():
#     print(f"{key}:")
#     for item in value:
#         print(item)
#     print()

# # Access the second product
# second_product = data_dict['Products'][1]  # Index 1 corresponds to the second element
# print("The second product is:", second_product)