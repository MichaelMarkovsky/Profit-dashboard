#for EXCEL
import os
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook


# import required module for DATABASE
import sqlite3





def create_excel():
    try:
        # Try to load the workbook if it exists
        workbook = openpyxl.load_workbook("./ExcelData.xlsx")
    except FileNotFoundError:
        # Create a new workbook and select the active sheet if the file does not exist
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sales"
    else:
        # Select the existing sheet
        worksheet = workbook['Sales']

    # Write data to the worksheet
    worksheet['A1'] = 'Products'
    worksheet['B1'] = 'Category'
    worksheet['C1'] = 'Date'
    worksheet['D1'] = 'Time of subscription'
    worksheet['E1'] = 'Price in shop'
    worksheet['F1'] = 'Price of bought'
    worksheet['G1'] = 'Profit'

    # Define fill color (example: solid fill with color yellow)
    fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    # Define font color (example: white font)
    white_font = Font(color="FFFFFF")

    # Apply the fill to the header cells
    cell_ids = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1']
    for cell_id in cell_ids:
        worksheet[cell_id].fill = fill
        worksheet[cell_id].font = white_font

    # Save the workbook
    workbook.save("./ExcelData.xlsx")



def create_database():

    def create_table():
        table = """
            CREATE TABLE IF NOT EXISTS products (
                name TEXT PRIMARY KEY,
                timesubscription TEXT NOT NULL,
                priceinshop TEXT NOT NULL,
                priceofbought TEXT NOT NULL,
                profit TEXT NOT NULL
            );
            """
        return table
    
   # Define the database file name
    db_name = 'products.db'

    # Check if the database file already exists
    if not os.path.exists(db_name):
        # Create a new database connection
        conn = sqlite3.connect(db_name)
        print(f"Database '{db_name}' created successfully.")
        
        # #create table
        # conn.execute(create_table())
        # # Commit the changes
        # conn.commit()

        # Close the connection
        conn.close()
    else:
        print(f"Database '{db_name}' already exists.")

  
create_database()